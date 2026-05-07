"""
Rutas para exportación de informes (Cartón dosimétrico e Informe final).

Para cada documento se exponen dos endpoints de descarga:
    - *_xlsx : devuelve el archivo Excel editable.
    - *_pdf  : devuelve el PDF generado a partir del MISMO Excel ya completado,
               de forma que ambos archivos tengan exactamente la misma información.
"""
import os
import io
import re
import json
from datetime import datetime
from flask import Blueprint, request, send_file
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from app.utils.file_handlers import (
    xlsx_to_pdf,
    pdf_to_png,
    insert_png_into_excel,
    write_to_excel_cell,
    safe_cell_write,
)
from app.utils.helpers import parse_patient_name, round_2_decimals
from config.settings import TEMPLATE_CARTON, TEMPLATE_INFORME

bp = Blueprint('export', __name__)


_MESES_ES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
]


def _format_fechas_es(date_strings):
    """
    Recibe strings 'YYYY-MM-DD' y devuelve la lista combinada en español.
    Ej: '20, 22, 27 y 29 de Octubre de 2020' (mismo mes/año)
        '20 de Octubre y 3 de Noviembre de 2020' (distinto mes)
    """
    parsed = []
    for s in date_strings:
        s = (s or "").strip()
        if not s:
            continue
        try:
            parsed.append(datetime.strptime(s, "%Y-%m-%d"))
        except ValueError:
            pass

    if not parsed:
        return ""

    parsed.sort()

    mismo_mes_anio = all(
        d.month == parsed[0].month and d.year == parsed[0].year
        for d in parsed
    )

    if mismo_mes_anio:
        mes = _MESES_ES[parsed[0].month - 1].capitalize()
        anio = parsed[0].year
        dias = [str(d.day) for d in parsed]
        if len(dias) == 1:
            return f"{dias[0]} de {mes} de {anio}"
        return f"{', '.join(dias[:-1])} y {dias[-1]} de {mes} de {anio}"

    partes = [
        f"{d.day} de {_MESES_ES[d.month - 1].capitalize()} de {d.year}"
        for d in parsed
    ]
    if len(partes) == 1:
        return partes[0]
    return f"{', '.join(partes[:-1])} y {partes[-1]}"


# ---------------------------------------------------------------------------
# Helpers comunes
# ---------------------------------------------------------------------------

def _parse_payload():
    """Lee y deserializa el JSON del campo 'payload' del form."""
    payload = request.form.get("payload", "")
    if not payload:
        return None, ("Sin datos para exportar", 400)
    try:
        return json.loads(payload), None
    except Exception as e:
        return None, (f"Payload inválido: {str(e)}", 400)


def _safe_filename_part(value):
    """Normaliza un string para que sea seguro como parte de un nombre de archivo."""
    if value is None:
        return ""
    cleaned = re.sub(r'[^\w\-]+', '_', str(value).strip(), flags=re.UNICODE)
    return cleaned.strip('_')


def _build_filename(prefix, patient_id, patient_name, ext):
    """
    Construye un nombre de archivo ordenado:
        Prefijo_<paciente>_<YYYYMMDD>.<ext>
    Usa patient_id si está disponible, si no patient_name. Si no hay ninguno,
    usa "paciente".
    """
    pid = _safe_filename_part(patient_id)
    pname = _safe_filename_part(patient_name)
    who = pid or pname or "paciente"
    fecha = datetime.today().strftime("%Y%m%d")
    return f"{prefix}_{who}_{fecha}.{ext}"


def _libreoffice_error_response(err):
    """
    Respuesta amigable cuando falla la conversión a PDF
    (típicamente porque LibreOffice no está instalado).
    """
    msg = (
        "<h2>No se pudo generar el PDF</h2>"
        "<p>La conversión de Excel a PDF requiere LibreOffice instalado en el "
        "servidor (modo <code>headless</code>).</p>"
        "<p>Mientras tanto, podés descargar la versión Excel editable y "
        "convertirla manualmente, o instalar LibreOffice desde "
        "<a href='https://www.libreoffice.org/'>libreoffice.org</a>.</p>"
        f"<p><strong>Detalle técnico:</strong> {str(err)}</p>"
    )
    return msg, 500


def _send_xlsx(excel_bytes, filename):
    return send_file(
        io.BytesIO(excel_bytes),
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _send_pdf(pdf_bytes, filename):
    return send_file(
        pdf_bytes,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )


# ---------------------------------------------------------------------------
# Cartón dosimétrico — construcción del Excel
# ---------------------------------------------------------------------------

def _build_carton_xlsx(data):
    """
    Completa la plantilla del cartón dosimétrico con los datos del paciente
    y devuelve los bytes del Excel resultante.
    """
    patient_name = data.get("patient_name") or ""
    patient_id = data.get("patient_id") or ""
    fx_rt = int(data.get("fx_rt") or 0)
    summary = data.get("summary") or []
    ebrt = data.get("ebrt") or []
    hdr_fractions = data.get("hdr_fractions") or []

    if not os.path.exists(TEMPLATE_CARTON):
        raise FileNotFoundError(
            f"Plantilla del cartón no encontrada en: {TEMPLATE_CARTON}"
        )

    wb = load_workbook(TEMPLATE_CARTON)
    ws = wb["Hoja1 (2)"] if "Hoja1 (2)" in wb.sheetnames else wb.active

    apellido, nombre = parse_patient_name(patient_name)

    write_to_excel_cell(wb, ws.title, "C8", (apellido or patient_name).upper(), 'left')
    write_to_excel_cell(wb, ws.title, "C9", (nombre or "").upper(), 'left')
    write_to_excel_cell(wb, ws.title, "H3", patient_id, 'center')

    # === Tratamiento de RT Externa ===
    write_to_excel_cell(wb, ws.title, "C12", round(fx_rt * 2, 2), 'center')
    write_to_excel_cell(wb, ws.title, "C13", fx_rt, 'center')

    ebrt_map = {}
    for row in ebrt:
        roi = (row.get("roi") or "").upper()
        if roi:
            ebrt_map[roi] = row

    ctv_row = ebrt_map.get("CTV")
    if ctv_row:
        dose_ctv = round_2_decimals(ctv_row.get("D_ext"))
        write_to_excel_cell(wb, ws.title, "C14", dose_ctv, 'center')

    oar_rows = [
        ("RECTO", 13),
        ("VEJIGA", 14),
        ("SIGMOIDE", 15),
        ("INTESTINO", 16),
    ]
    for roi_key, row_idx in oar_rows:
        row = ebrt_map.get(roi_key)
        if row:
            dose_val = round_2_decimals(row.get("D_ext"))
            safe_cell_write(ws, row_idx, 9, dose_val, 'center')

    # === Tabla HDR (EQD2 por sesión) ===
    row_map_hdr = {
        "CTV": 24,
        "Recto": 25,
        "Vejiga": 26,
        "Sigmoide": 27,
        "Intestino": 28,
    }

    hdr_map = {(x["roi"] or "").upper(): x for x in hdr_fractions}

    def match_roi(roi_excel):
        roi_excel = roi_excel.upper()
        for roi_hdr, item in hdr_map.items():
            if roi_excel == "CTV" and "CTV" in roi_hdr:
                return item
            if roi_excel != "CTV" and roi_excel in roi_hdr:
                return item
        return None

    session_cols = [3, 4, 6, 8]

    for roi_excel, row_idx in row_map_hdr.items():
        item = match_roi(roi_excel)
        doses = [round_2_decimals(v) for v in item["doses"]] if item else []
        for j in range(4):
            col = session_cols[j]
            cell = ws.cell(row=row_idx, column=col)

            from openpyxl.cell.cell import MergedCell
            if isinstance(cell, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        cell = ws.cell(
                            row=merged_range.min_row,
                            column=merged_range.min_col
                        )
                        break

            if j < len(doses) and doses[j] is not None:
                cell.value = doses[j]
            else:
                cell.value = "-"

            cell.alignment = Alignment(horizontal='center')

    # === Registro de dosis total (EQD2) ===
    row_map_total = {
        "CTV": 35,
        "RECTO": 36,
        "VEJIGA": 37,
        "SIGMOIDE": 38,
        "INTESTINO": 39,
    }

    for item in summary:
        roi_raw = (item.get("roi") or "").upper()
        roi_key = roi_raw.replace(" (D90)", "")
        row_idx = row_map_total.get(roi_key)
        if not row_idx:
            continue

        safe_cell_write(ws, row_idx, 3, round_2_decimals(item.get("eqd2_ebrt")), 'center')
        safe_cell_write(ws, row_idx, 4, round_2_decimals(item.get("eqd2_hdr")), 'center')
        safe_cell_write(ws, row_idx, 7, round_2_decimals(item.get("eqd2_total")), 'center')

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    return excel_buffer.getvalue()


# ---------------------------------------------------------------------------
# Informe final — construcción del Excel
# ---------------------------------------------------------------------------

def _build_informe_xlsx(data, plan_pdf_storage=None, form=None):
    """
    Completa la plantilla del informe final y, opcionalmente, le inserta la
    imagen del plan (PDF subido por el usuario).

    `form` es un dict-like (típicamente request.form) con los campos extras
    completados por el usuario en la pantalla:
        inf_diagnostico, inf_braqui, inf_aplicador,
        inf_sesiones, inf_dosis_gy,
        inf_fecha_1..inf_fecha_4,
        inf_dur_num, inf_dur_unit
    """
    patient_name = data.get("patient_name") or ""
    summary = data.get("summary") or []
    form = form or {}

    if not os.path.exists(TEMPLATE_INFORME):
        raise FileNotFoundError(
            f"Plantilla del informe no encontrada en: {TEMPLATE_INFORME}"
        )

    wb = load_workbook(TEMPLATE_INFORME)
    ws = wb.worksheets[0]

    # El título "Resumen dosimétrico..." en B29:I29 tiene wrapText+vertical=bottom;
    # con altura de 1 línea LibreOffice solo muestra la última línea al exportar a PDF.
    ws.row_dimensions[29].height = 30

    write_to_excel_cell(wb, ws.title, "G7", datetime.today().strftime("%d/%m/%Y"))
    write_to_excel_cell(wb, ws.title, "G12", patient_name)

    row_map = {
        "CTV": 32,
        "RECTO": 33,
        "VEJIGA": 34,
        "SIGMOIDE": 35,
        "INTESTINO": 36,
    }

    for item in summary:
        roi_raw = (item.get("roi") or "").upper()
        roi_key = roi_raw.replace(" (D90)", "")
        row = row_map.get(roi_key)
        if not row:
            continue

        safe_cell_write(ws, row, 4, round_2_decimals(item.get("eqd2_ebrt")), 'center')
        safe_cell_write(ws, row, 6, round_2_decimals(item.get("eqd2_hdr")), 'center')
        safe_cell_write(ws, row, 8, round_2_decimals(item.get("eqd2_total")), 'center')

    # === Campos adicionales del informe (autocompletado) ===
    inf_diagnostico = (form.get("inf_diagnostico") or "").strip()
    inf_braqui = (form.get("inf_braqui") or "").strip()
    inf_aplicador = (form.get("inf_aplicador") or "").strip()
    inf_sesiones = (form.get("inf_sesiones") or "").strip()
    inf_dosis_gy = (form.get("inf_dosis_gy") or "").strip()

    if inf_diagnostico:
        write_to_excel_cell(wb, ws.title, "E13", inf_diagnostico, 'left')
    if inf_braqui:
        write_to_excel_cell(wb, ws.title, "G15", inf_braqui, 'left')
    if inf_aplicador:
        write_to_excel_cell(wb, ws.title, "C21", inf_aplicador, 'left')

    if inf_sesiones and inf_dosis_gy:
        try:
            n_ses = int(inf_sesiones)
            d_gy = float(inf_dosis_gy)
            write_to_excel_cell(
                wb, ws.title, "D24",
                f"{n_ses} sesiones de {d_gy:g} Gy",
                'center'
            )
        except ValueError:
            pass

    # Fechas de tratamiento (E26)
    raw_fechas = [form.get(f"inf_fecha_{i}", "") for i in range(1, 5)]
    fechas_str = _format_fechas_es(raw_fechas)
    if fechas_str:
        write_to_excel_cell(wb, ws.title, "E26", fechas_str, 'left')

    # Duración del tratamiento (E27)
    inf_dur_num = (form.get("inf_dur_num") or "").strip()
    inf_dur_unit = (form.get("inf_dur_unit") or "semanas").strip()
    if inf_dur_num:
        try:
            n_dur = int(inf_dur_num)
            write_to_excel_cell(
                wb, ws.title, "E27",
                f"{n_dur} {inf_dur_unit}",
                'left'
            )
        except ValueError:
            pass

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()

    if plan_pdf_storage and getattr(plan_pdf_storage, "filename", ""):
        try:
            pdf_bytes = plan_pdf_storage.read()
            png_bytes = pdf_to_png(pdf_bytes, rotation=90, dpi=150)
            excel_bytes = insert_png_into_excel(
                excel_bytes,
                png_bytes,
                sheet_name="IMAGEN",
                cell="B5",
                scale=0.35
            )
        except Exception as e:
            # Si falla la inserción de la imagen, seguimos con el Excel sin imagen.
            print(f"Advertencia: No se pudo insertar imagen del plan: {str(e)}")

    return excel_bytes


# ---------------------------------------------------------------------------
# Endpoints — Cartón dosimétrico
# ---------------------------------------------------------------------------

@bp.route("/export_carton_xlsx", methods=["POST"])
def export_carton_xlsx():
    """Descarga el cartón dosimétrico como Excel editable (.xlsx)."""
    data, err = _parse_payload()
    if err:
        return err

    try:
        excel_bytes = _build_carton_xlsx(data)
    except FileNotFoundError as e:
        return str(e), 500
    except Exception as e:
        return f"Error al preparar el Excel del cartón: {str(e)}", 500

    filename = _build_filename(
        "Carton", data.get("patient_id"), data.get("patient_name"), "xlsx"
    )
    return _send_xlsx(excel_bytes, filename)


@bp.route("/export_carton_pdf", methods=["POST"])
def export_carton_pdf():
    """
    Descarga el cartón dosimétrico como PDF.
    El PDF se genera a partir del MISMO Excel completado, para garantizar
    que ambas descargas contengan la misma información.
    """
    data, err = _parse_payload()
    if err:
        return err

    try:
        excel_bytes = _build_carton_xlsx(data)
    except FileNotFoundError as e:
        return str(e), 500
    except Exception as e:
        return f"Error al preparar el Excel del cartón: {str(e)}", 500

    try:
        pdf_bytes = xlsx_to_pdf(excel_bytes)
    except RuntimeError as e:
        return _libreoffice_error_response(e)
    except Exception as e:
        return f"Error al convertir a PDF: {str(e)}", 500

    filename = _build_filename(
        "Carton", data.get("patient_id"), data.get("patient_name"), "pdf"
    )
    return _send_pdf(pdf_bytes, filename)


# Alias retrocompatible: el endpoint anterior devolvía siempre PDF.
@bp.route("/export_carton", methods=["POST"])
def export_carton():
    return export_carton_pdf()


# ---------------------------------------------------------------------------
# Endpoints — Informe final
# ---------------------------------------------------------------------------

@bp.route("/export_informe_xlsx", methods=["POST"])
def export_informe_xlsx():
    """Descarga el informe final como Excel editable (.xlsx)."""
    data, err = _parse_payload()
    if err:
        return err

    plan_pdf = request.files.get("plan_pdf")
    try:
        excel_bytes = _build_informe_xlsx(data, plan_pdf, request.form)
    except FileNotFoundError as e:
        return str(e), 500
    except Exception as e:
        return f"Error al preparar el Excel del informe: {str(e)}", 500

    filename = _build_filename(
        "Informe_final", data.get("patient_id"), data.get("patient_name"), "xlsx"
    )
    return _send_xlsx(excel_bytes, filename)


@bp.route("/export_informe_pdf", methods=["POST"])
def export_informe_pdf():
    """
    Descarga el informe final como PDF, generado a partir del mismo Excel
    completado (con la imagen del plan si se adjuntó).
    """
    data, err = _parse_payload()
    if err:
        return err

    plan_pdf = request.files.get("plan_pdf")
    try:
        excel_bytes = _build_informe_xlsx(data, plan_pdf, request.form)
    except FileNotFoundError as e:
        return str(e), 500
    except Exception as e:
        return f"Error al preparar el Excel del informe: {str(e)}", 500

    try:
        pdf_bytes = xlsx_to_pdf(excel_bytes)
    except RuntimeError as e:
        return _libreoffice_error_response(e)
    except Exception as e:
        return f"Error al convertir a PDF: {str(e)}", 500

    filename = _build_filename(
        "Informe_final", data.get("patient_id"), data.get("patient_name"), "pdf"
    )
    return _send_pdf(pdf_bytes, filename)


# Alias retrocompatible: el endpoint anterior devolvía siempre Excel.
@bp.route("/export_informe", methods=["POST"])
def export_informe():
    return export_informe_xlsx()
